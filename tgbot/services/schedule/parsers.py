"""Optimized and refactored schedule parsers with common utilities."""

import calendar
import logging
import re
from abc import ABC, abstractmethod
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pytz
from openpyxl import load_workbook
from pandas import DataFrame
from stp_database import Employee, MainRequestsRepo

from ...misc.helpers import format_fullname
from . import DutyInfo, HeadInfo
from .analyzers import ScheduleAnalyzer
from .formatters import ScheduleFormatter
from .managers import MonthManager, ScheduleFileManager
from .models import GroupMemberInfo

logger = logging.getLogger(__name__)


class CommonUtils:
    """Common utility functions used across all parsers."""

    @staticmethod
    def get_cell_value(df: pd.DataFrame, row: int, col: int) -> str:
        """Safely get cell value from DataFrame."""
        try:
            if row < len(df) and col < len(df.columns):
                value = df.iloc[row, col]
                return str(value) if pd.notna(value) else ""
            return ""
        except (IndexError, TypeError):
            return ""

    @staticmethod
    def is_valid_name(name_cell: str) -> bool:
        """Check if cell contains a valid full name."""
        if not name_cell or name_cell.strip() in ["", "nan", "None"]:
            return False
        parts = name_cell.strip().split()
        return len(parts) >= 2

    @staticmethod
    def names_match(name1: str, name2: str) -> bool:
        """Check if two names match (considering minor differences)."""
        if not name1 or not name2:
            return False

        name1_clean = name1.strip()
        name2_clean = name2.strip()

        if name1_clean == name2_clean:
            return True

        parts1 = name1_clean.split()
        parts2 = name2_clean.split()

        if len(parts1) >= 2 and len(parts2) >= 2:
            return parts1[0] == parts2[0] and parts1[1] == parts2[1]

        return False

    @staticmethod
    def is_time_format(text: str) -> bool:
        """Check if text contains time format (HH:MM-HH:MM)."""
        if not text:
            return False
        time_pattern = r"\d{1,2}:\d{2}-\d{1,2}:\d{2}"
        return bool(re.search(time_pattern, text.strip()))

    @staticmethod
    def parse_time_range(time_str: str) -> Tuple[int, int]:
        """Parse time range string into start and end minutes."""
        try:
            if "-" not in time_str:
                return 0, 0

            start_time, end_time = time_str.split("-")
            start_parts = start_time.strip().split(":")
            end_parts = end_time.strip().split(":")

            start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
            end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])

            # Handle overnight shifts
            if end_minutes < start_minutes:
                end_minutes += 24 * 60

            return start_minutes, end_minutes

        except (ValueError, IndexError):
            return 0, 0


class TimeUtils:
    """Utilities for time parsing and sorting."""

    @staticmethod
    def extract_start_time(working_hours: str) -> str:
        """Extract start time from working hours string."""
        if not working_hours or working_hours == "Не указано":
            return "Не указано"

        time_pattern = r"(\d{1,2}:\d{2})"
        match = re.search(time_pattern, working_hours)

        if match:
            return match.group(1)

        return "Не указано"

    @staticmethod
    def parse_time_for_sorting(time_str: str) -> Tuple[int, int]:
        """Parse time string for sorting purposes."""
        if not time_str or time_str == "Не указано":
            return 99, 0  # Put "not specified" at the end

        try:
            hour, minute = time_str.split(":")
            return int(hour), int(minute)
        except (ValueError, IndexError):
            return 99, 0


class DateColumnFinder:
    """Utility class for finding date columns in DataFrames."""

    @staticmethod
    def find_date_column(
        df: pd.DataFrame, target_date: datetime, search_rows: int = 5
    ) -> Optional[int]:
        """Find column for target date by first locating the correct month section."""
        target_day = target_date.day
        target_month = target_date.month

        # Russian month names
        month_names = {
            1: "ЯНВАРЬ",
            2: "ФЕВРАЛЬ",
            3: "МАРТ",
            4: "АПРЕЛЬ",
            5: "МАЙ",
            6: "ИЮНЬ",
            7: "ИЮЛЬ",
            8: "АВГУСТ",
            9: "СЕНТЯБРЬ",
            10: "ОКТЯБРЬ",
            11: "НОЯБРЬ",
            12: "ДЕКАБРЬ",
        }

        target_month_name = month_names[target_month]
        logger.debug(f"Searching for day {target_day} in month '{target_month_name}'")

        # Step 1: Find the month section
        month_start_col = None

        # Look for the month header
        for row_idx in range(min(3, len(df))):
            for col_idx in range(len(df.columns)):
                cell_value = CommonUtils.get_cell_value(df, row_idx, col_idx)

                if target_month_name in cell_value.upper():
                    month_start_col = col_idx
                    logger.debug(
                        f"Found month '{target_month_name}' starting at column {col_idx}"
                    )
                    break
            if month_start_col is not None:
                break

        if month_start_col is None:
            logger.warning(f"Month '{target_month_name}' not found in headers")
            # Fallback to old method
            return DateColumnFinder._find_date_column_fallback(
                df, target_date, search_rows
            )

        # Step 2: Determine the end of this month's section
        # Look for the next month header or end of data
        month_end_col = len(df.columns) - 1  # Default to end of sheet

        for next_month in range(target_month + 1, 13):  # Check subsequent months
            next_month_name = month_names[next_month]
            for row_idx in range(min(3, len(df))):
                for col_idx in range(month_start_col + 1, len(df.columns)):
                    cell_value = CommonUtils.get_cell_value(df, row_idx, col_idx)
                    if next_month_name in cell_value.upper():
                        month_end_col = col_idx - 1
                        logger.debug(
                            f"Month section ends at column {month_end_col} (before {next_month_name})"
                        )
                        break
                if month_end_col < len(df.columns) - 1:
                    break
            if month_end_col < len(df.columns) - 1:
                break

        logger.debug(
            f"Searching for day {target_day} in columns {month_start_col} to {month_end_col}"
        )

        # Step 3: Look for the target day within this month's section
        for row_idx in range(min(search_rows, len(df))):
            for col_idx in range(month_start_col, month_end_col + 1):
                cell_value = CommonUtils.get_cell_value(df, row_idx, col_idx)

                if not cell_value:
                    continue

                cell_value = cell_value.strip()

                # Look for exact day pattern like "28Чт"
                day_pattern = r"^(\d{1,2})[А-Яа-я]{0,2}$"
                match = re.search(day_pattern, cell_value)

                if match and int(match.group(1)) == target_day:
                    logger.debug(
                        f"Found day {target_day} at row {row_idx}, col {col_idx} in {target_month_name}: '{cell_value}'"
                    )
                    return col_idx

                # Also check for simple day number
                if cell_value == str(target_day):
                    logger.debug(
                        f"Found simple day {target_day} at row {row_idx}, col {col_idx} in {target_month_name}"
                    )
                    return col_idx

        logger.warning(f"Day {target_day} not found in {target_month_name} section")
        return None

    @staticmethod
    def _find_date_column_fallback(
        df: pd.DataFrame, target_date: datetime, search_rows: int = 5
    ) -> Optional[int]:
        """Fallback method - original logic for when month-based search fails."""
        target_day = target_date.day

        for row_idx in range(min(search_rows, len(df))):
            for col_idx in range(len(df.columns)):
                cell_value = CommonUtils.get_cell_value(df, row_idx, col_idx)

                if not cell_value:
                    continue

                # Pattern for day with letters (like "28Чт")
                day_pattern = r"^(\d{1,2})[А-Яа-я]{1,2}$"
                match = re.search(day_pattern, cell_value.strip())

                if match and int(match.group(1)) == target_day:
                    logger.debug(f"Fallback: Found date column {target_day}: {col_idx}")
                    return col_idx

        return None


class BaseExcelParser(ABC):
    """Abstract base class for Excel parsers."""

    def __init__(self, uploads_folder: str = "uploads"):
        self.file_manager = ScheduleFileManager(uploads_folder)
        self.utils = CommonUtils()
        self.date_finder = DateColumnFinder()
        self.time_utils = TimeUtils()

    @staticmethod
    def read_excel_file(
        file_path: Path, sheet_name: str = "ГРАФИК"
    ) -> Optional[DataFrame]:
        """Read Excel file and return DataFrame."""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            logger.debug(f"Successfully read sheet: {sheet_name}")
            return df
        except Exception as e:
            logger.debug(f"Failed to read sheet '{sheet_name}': {e}")
            return None

    def find_user_row(
        self, df: pd.DataFrame, fullname: str, search_cols: int = 3
    ) -> Optional[int]:
        """Find row containing user's full name."""
        for row_idx in range(len(df)):
            for col_idx in range(min(search_cols, len(df.columns))):
                cell_value = self.utils.get_cell_value(df, row_idx, col_idx)

                if fullname in cell_value:
                    logger.debug(f"User '{fullname}' found in row {row_idx}")
                    return row_idx

        return None

    @abstractmethod
    def format_schedule(self, data: List, date: datetime) -> str:
        """Format schedule data for display."""
        pass


class MonthlyScheduleParser(BaseExcelParser, ABC):
    """Parser for monthly schedules."""

    @staticmethod
    def find_month_columns(df: pd.DataFrame, month: str) -> Tuple[int, int]:
        """Find start and end columns for specified month."""
        month = MonthManager.normalize_month(month)

        def find_month_index(
            target_month: str, target_first_col: int = 0
        ) -> Optional[int]:
            for col_idx in range(target_first_col, len(df.columns)):
                # Check in headers
                col_name = (
                    str(df.columns[col_idx]).upper() if df.columns[col_idx] else ""
                )
                if target_month in col_name:
                    return col_idx

                # Check in first few rows
                for row_idx in range(min(5, len(df))):
                    val = df.iat[row_idx, col_idx]
                    if isinstance(val, str) and target_month in val.upper():
                        return col_idx
            return None

        start_column = find_month_index(month)
        if start_column is None:
            raise ValueError(f"Month {month} not found in schedule")

        # Find end column (before next month)
        end_column = len(df.columns) - 1
        for m in MonthManager.MONTHS_ORDER:
            if m != month:
                next_month_col = find_month_index(m, start_column + 1)
                if next_month_col is not None:
                    end_column = next_month_col - 1
                    break

        logger.debug(f"Month '{month}' found in columns {start_column}-{end_column}")
        return start_column, end_column

    def find_day_headers(
        self, df: pd.DataFrame, start_column: int, end_column: int
    ) -> Dict[int, str]:
        """Find day headers in specified column range."""
        day_headers = {}

        for row_idx in range(min(5, len(df))):
            for col_idx in range(start_column, end_column + 1):
                cell_value = self.utils.get_cell_value(df, row_idx, col_idx)

                # Pattern for day with letters
                day_pattern = r"(\d{1,2})([А-Яа-я]{1,2})"
                match = re.search(day_pattern, cell_value)

                if match:
                    day_num = match.group(1)
                    day_name = match.group(2)
                    day_headers[col_idx] = f"{day_num} ({day_name})"
                elif (
                    cell_value.strip().isdigit() and 1 <= int(cell_value.strip()) <= 31
                ):
                    day_headers[col_idx] = cell_value.strip()

        logger.debug(f"Found {len(day_headers)} days in headers")
        return day_headers


class ScheduleParser(MonthlyScheduleParser):
    """Main schedule parser class."""

    def __init__(self, uploads_folder: str = "uploads"):
        super().__init__(uploads_folder)
        self.analyzer = ScheduleAnalyzer()
        self.formatter = ScheduleFormatter()

    def get_user_schedule(
        self, fullname: str, month: str, division: str
    ) -> Dict[str, str]:
        """Get user's schedule for specified month."""
        try:
            schedule_file = self.file_manager.find_schedule_file(division)
            if not schedule_file:
                raise FileNotFoundError(f"Файл графика для {division} не найден")

            df = self.read_excel_file(schedule_file)
            if df is None:
                raise ValueError("Не удалось прочитать файл графика")

            start_column, end_column = self.find_month_columns(df, month)
            day_headers = self.find_day_headers(df, start_column, end_column)

            user_row_idx = self.find_user_row(df, fullname)
            if user_row_idx is None:
                raise ValueError(f"Сотрудник {fullname} не найден в графике")

            schedule = {}
            for col_idx in range(start_column, end_column + 1):
                if col_idx in day_headers:
                    day = day_headers[col_idx]
                    schedule_value = self.utils.get_cell_value(
                        df, user_row_idx, col_idx
                    ).strip()

                    if schedule_value.lower() in ["nan", "none", "", "0", "0.0"]:
                        schedule_value = "Не указано"

                    schedule[day] = schedule_value

            logger.info(f"Found {len(schedule)} days for {fullname} in {month}")
            return schedule

        except Exception as e:
            logger.error(f"Error getting schedule: {e}")
            raise

    def get_user_schedule_with_additional_shifts(
        self, fullname: str, month: str, division: str
    ) -> Tuple[Dict[str, str], Dict[str, str]]:
        """Get user's schedule with additional shifts detected by color #cc99ff."""
        try:
            schedule_file = self.file_manager.find_schedule_file(division)
            if not schedule_file:
                raise FileNotFoundError(f"Файл графика для {division} не найден")

            # Load with openpyxl to access cell formatting
            wb = load_workbook(schedule_file, data_only=False)
            ws = wb["ГРАФИК"] if "ГРАФИК" in wb.sheetnames else wb.active

            # Also load with pandas for easier data access
            df = self.read_excel_file(schedule_file)
            if df is None:
                raise ValueError("Не удалось прочитать файл графика")

            start_column, end_column = self.find_month_columns(df, month)
            day_headers = self.find_day_headers(df, start_column, end_column)

            user_row_idx = self.find_user_row(df, fullname)
            if user_row_idx is None:
                raise ValueError(f"Сотрудник {fullname} не найден в графике")

            schedule = {}
            additional_shifts = {}

            for col_idx in range(start_column, end_column + 1):
                if col_idx in day_headers:
                    day = day_headers[col_idx]
                    schedule_value = self.utils.get_cell_value(
                        df, user_row_idx, col_idx
                    ).strip()

                    if schedule_value.lower() in ["nan", "none", "", "0", "0.0"]:
                        schedule_value = "Не указано"

                    # Check cell color in openpyxl (1-indexed)
                    cell = ws.cell(row=user_row_idx + 1, column=col_idx + 1)
                    is_additional_shift = self._is_additional_shift_color(cell)

                    if is_additional_shift and schedule_value not in [
                        "Не указано",
                        "В",
                        "О",
                        "0",
                        "0.0",
                    ]:
                        additional_shifts[day] = schedule_value
                    else:
                        schedule[day] = schedule_value

            logger.debug(
                f"Found {len(schedule)} regular days and {len(additional_shifts)} additional shifts for {fullname} in {month}"
            )
            return schedule, additional_shifts

        except Exception as e:
            logger.error(f"Error getting schedule with additional shifts: {e}")
            raise

    @staticmethod
    def _is_additional_shift_color(cell) -> bool:
        """Check if cell has the additional shift color #cc99ff."""
        try:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color

                # Convert color to hex if it's a Color object
                if hasattr(color, "rgb") and color.rgb:
                    hex_color = color.rgb
                    if isinstance(hex_color, str) and len(hex_color) >= 6:
                        # Remove alpha channel if present (ARGB -> RGB)
                        if len(hex_color) == 8:
                            hex_color = hex_color[2:]
                        # Check if it matches #cc99ff (case insensitive)
                        return hex_color.lower() == "cc99ff"

                # Also check indexed colors if available
                if hasattr(color, "indexed") and color.indexed is not None:
                    # This would require a color palette lookup
                    # For now, we'll rely on RGB values
                    pass

            return False
        except Exception as e:
            logger.debug(f"Error checking cell color: {e}")
            return False

    async def get_user_schedule_with_duties(
        self, fullname: str, month: str, division: str, stp_repo=None
    ) -> Dict[str, tuple[str, Optional[str]]]:
        """Get user's schedule with duty information for specified month."""
        try:
            # Get regular schedule data
            schedule_data = self.get_user_schedule(fullname, month, division)

            if not schedule_data or not stp_repo:
                return {
                    day: (schedule, None) for day, schedule in schedule_data.items()
                }

            # Get duty parser to check for duty information
            duty_parser = DutyScheduleParser()

            # OPTIMIZATION: Get all duties for the month at once instead of day by day
            current_year = datetime.now().year
            month_num = MonthManager.get_month_number(month)

            try:
                # Create a date object for the first day of the month to get month duties
                first_day_of_month = datetime(current_year, month_num, 1)

                # Get all duties for the entire month at once
                month_duties = await duty_parser.get_duties_for_month(
                    first_day_of_month, division, stp_repo
                )

                logger.debug(
                    f"Retrieved duties for {len(month_duties)} days in month {month}"
                )

            except Exception as e:
                logger.warning(
                    f"Failed to get month duties, falling back to individual day queries: {e}"
                )
                month_duties = {}

            # Create result with duty information
            schedule_with_duties = {}

            for day, schedule in schedule_data.items():
                duty_info = None

                # Try to parse the date and check for duty information
                try:
                    # Extract day number from day string (e.g., "1 (Пн)" -> 1)
                    import re

                    day_match = re.search(r"(\d+)", day)
                    if day_match:
                        day_num = int(day_match.group(1))

                        # Check if we have month duties data
                        if month_duties and day_num in month_duties:
                            # Use optimized month data
                            duties = month_duties[day_num]
                        else:
                            # Fallback to individual day query
                            try:
                                date = datetime(current_year, month_num, day_num)
                                duties = await duty_parser.get_duties_for_date(
                                    date, division, stp_repo
                                )
                            except ValueError:
                                # Invalid date, skip duty check
                                duties = []

                        # Check if user is on duty
                        for duty in duties:
                            if self.utils.names_match(fullname, duty.name):
                                duty_info = f"{duty.schedule} {duty.shift_type}"
                                break

                except Exception as e:
                    logger.debug(f"Error checking duty for {fullname} on {day}: {e}")

                schedule_with_duties[day] = (schedule, duty_info)

            return schedule_with_duties

        except Exception as e:
            logger.error(f"Error getting schedule with duties: {e}")
            # Fallback to regular schedule without duties
            schedule_data = self.get_user_schedule(fullname, month, division)
            return {day: (schedule, None) for day, schedule in schedule_data.items()}

    def get_user_schedule_formatted(
        self,
        fullname: str,
        month: str,
        division: str,
        compact: bool = False,
    ) -> str:
        """Get formatted user schedule."""
        try:
            schedule_data = self.get_user_schedule(fullname, month, division)

            if not schedule_data:
                return f"❌ Schedule for <b>{fullname}</b> in {month} not found"

            analysis = self.analyzer.analyze_schedule(schedule_data)

            if compact:
                return self.formatter.format_compact(month, *analysis)
            else:
                return self.formatter.format_detailed(month, *analysis)

        except Exception as e:
            logger.error(f"Schedule formatting error: {e}")
            return f"❌ <b>Ошибка графика:</b>\n<code>{e}</code>"

    async def get_user_schedule_formatted_with_duties(
        self,
        fullname: str,
        month: str,
        division: str,
        compact: bool = False,
        stp_repo=None,
    ) -> str:
        """Get formatted user schedule with duty information."""
        try:
            schedule_data_with_duties = await self.get_user_schedule_with_duties(
                fullname, month, division, stp_repo
            )

            if not schedule_data_with_duties:
                return f"❌ Schedule for <b>{fullname}</b> in {month} not found"

            # Extract regular schedule data for analysis
            schedule_data = {
                day: schedule
                for day, (schedule, _) in schedule_data_with_duties.items()
            }
            analysis = self.analyzer.analyze_schedule(schedule_data)

            if compact:
                # For compact view, use regular formatting without duties
                return self.formatter.format_compact(month, *analysis)
            else:
                return self.formatter.format_detailed_with_duties(
                    month, schedule_data_with_duties, *analysis
                )

        except Exception as e:
            logger.error(f"Schedule formatting error: {e}")
            return f"❌ <b>Ошибка графика:</b>\n<code>{e}</code>"

    def format_schedule(self, data: List, date: datetime) -> str:
        """Format schedule data for display."""
        # Implementation depends on data structure
        return ""


class BaseDutyParser(BaseExcelParser, ABC):
    """Base class for duty-related parsers."""

    @staticmethod
    def get_duty_sheet_name(date: datetime) -> str:
        """Generate sheet name for duty schedule."""
        month_names = [
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        ]
        month_name = month_names[date.month - 1]
        return f"Дежурство {month_name}"

    @staticmethod
    def parse_duty_entry(cell_value: str) -> Tuple[str, str]:
        """Parse duty entry to extract shift type and time."""
        if not cell_value or cell_value.strip() in ["", "nan", "None"]:
            return "", ""

        cell_value = cell_value.strip()

        if cell_value.startswith("П "):
            return "П", cell_value[2:].strip()
        elif cell_value.startswith("С "):
            return "С", cell_value[2:].strip()
        else:
            if re.search(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}", cell_value):
                return "", cell_value
            else:
                return "", cell_value


class DutyScheduleParser(BaseDutyParser):
    """Parser for duty schedules."""

    async def get_current_senior_duty(
        self, division: str, stp_repo: MainRequestsRepo
    ) -> Optional[DutyInfo]:
        """Get current senior duty for division based on current time."""
        yekaterinburg_tz = pytz.timezone("Asia/Yekaterinburg")
        date = datetime.now(yekaterinburg_tz)

        try:
            # Get all duties for today
            duties = await self.get_duties_for_date(date, division, stp_repo)

            # Filter for senior duties only
            senior_duties = [duty for duty in duties if duty.shift_type == "С"]

            if not senior_duties:
                return None

            # Find the current senior duty based on time
            for duty in senior_duties:
                if self.utils.is_time_format(duty.schedule):
                    start_minutes, end_minutes = self.utils.parse_time_range(
                        duty.schedule
                    )
                    current_datetime = date
                    current_time_minutes = (
                        current_datetime.hour * 60 + current_datetime.minute
                    )
                    # Check if current time is within duty hours
                    if start_minutes <= current_time_minutes <= end_minutes:
                        return duty

                    # Handle overnight shifts (end time is next day)
                    elif end_minutes > 24 * 60:  # Overnight shift
                        if (
                            current_time_minutes >= start_minutes
                            or current_time_minutes <= (end_minutes - 24 * 60)
                        ):
                            return duty

            # If no active duty found, return None (no fallback)
            return None

        except Exception as e:
            logger.error(f"Error getting current senior duty for {division}: {e}")
            return None

    async def get_current_helper_duty(
        self, division: str, stp_repo: MainRequestsRepo
    ) -> Optional[DutyInfo]:
        """Get current helper duty for division based on current time."""
        yekaterinburg_tz = pytz.timezone("Asia/Yekaterinburg")
        date = datetime.now(yekaterinburg_tz)

        try:
            # Get all duties for today
            duties = await self.get_duties_for_date(date, division, stp_repo)

            # Filter for helper duties only
            helper_duties = [duty for duty in duties if duty.shift_type == "П"]

            if not helper_duties:
                return None

            # Find the current helper duty based on time
            for duty in helper_duties:
                if self.utils.is_time_format(duty.schedule):
                    start_minutes, end_minutes = self.utils.parse_time_range(
                        duty.schedule
                    )
                    current_datetime = date.now()
                    current_time_minutes = (
                        current_datetime.hour * 60 + current_datetime.minute
                    )
                    # Check if current time is within duty hours
                    if start_minutes <= current_time_minutes <= end_minutes:
                        return duty

                    # Handle overnight shifts (end time is next day)
                    elif end_minutes > 24 * 60:  # Overnight shift
                        if (
                            current_time_minutes >= start_minutes
                            or current_time_minutes <= (end_minutes - 24 * 60)
                        ):
                            return duty

            # If no active duty found, return None (no fallback)
            return None

        except Exception as e:
            logger.error(f"Error getting current helper duty for {division}: {e}")
            return None

    async def get_duties_for_month(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> Dict[int, List[DutyInfo]]:
        """Get list of duty officers for entire month. Returns dict with day number as key."""
        try:
            # For НТП divisions, use separate seniority file
            if division in ["НТП", "НТП1", "НТП2"]:
                duty_file = self.file_manager.uploads_folder / "Старшинство_НТП.xlsx"
                if not duty_file.exists():
                    raise FileNotFoundError(
                        "Файл графика дежурных 'Старшинство_НТП.xlsx' не найден"
                    )
                schedule_file = duty_file
            # For НЦК division, use separate seniority file
            elif division == "НЦК":
                duty_file = self.file_manager.uploads_folder / "Старшинство_НЦК.xlsx"
                if not duty_file.exists():
                    raise FileNotFoundError(
                        "Файл графика дежурных 'Старшинство_НЦК.xlsx' не найден"
                    )
                schedule_file = duty_file
            else:
                schedule_file = self.file_manager.find_schedule_file(division)
                if not schedule_file:
                    raise FileNotFoundError(
                        f"Файл графика дежурных для {division} не найден"
                    )

            sheet_name = self.get_duty_sheet_name(date)

            try:
                df = self.read_excel_file(schedule_file, sheet_name)
            except Exception as e:
                logger.warning(f"Failed to read schedule with primary sheet name: {e}")

                # For НТП and НЦК divisions, try different sheet name patterns
                if division in ["НТП", "НТП1", "НТП2", "НЦК"]:
                    # Try just the month name
                    month_names = [
                        "Январь",
                        "Февраль",
                        "Март",
                        "Апрель",
                        "Май",
                        "Июнь",
                        "Июль",
                        "Август",
                        "Сентябрь",
                        "Октябрь",
                        "Ноябрь",
                        "Декабрь",
                    ]
                    month_name = month_names[date.month - 1]
                    try:
                        df = self.read_excel_file(schedule_file, month_name)
                        logger.debug(
                            f"Successfully read {division} duty sheet with name: {month_name}"
                        )
                    except Exception as e2:
                        logger.warning(
                            f"Failed to read {division} duty sheet with month name '{month_name}': {e2}"
                        )
                        df = None
                else:
                    df = None

            if df is None:
                logger.debug(
                    f"[График дежурных] Не удалось найти график дежурных на {date} для {division}"
                )
                raise ValueError("Не удалось найти график дежурных")

            # Find all date columns for the month
            month_duties = {}

            # Find all columns that might contain days from this month
            days_in_month = calendar.monthrange(date.year, date.month)[1]

            # Find date columns for all days in the month
            date_columns = {}
            for day in range(1, days_in_month + 1):
                try:
                    day_date = datetime(date.year, date.month, day)
                    date_col = self.date_finder.find_date_column(
                        df, day_date, search_rows=3
                    )
                    if date_col is not None:
                        date_columns[day] = date_col
                except ValueError:
                    # Invalid date, skip
                    continue

            logger.debug(
                f"Found date columns for {len(date_columns)} days in month {date.month}"
            )

            # Parse duties for all found dates at once
            for row_idx in range(len(df)):
                # Find name in first few columns
                name = ""
                for col_idx in range(min(3, len(df.columns))):
                    cell_value = self.utils.get_cell_value(df, row_idx, col_idx)

                    if (
                        len(cell_value.split()) >= 3
                        and re.search(r"[А-Яа-я]", cell_value)
                        and not re.search(r"\d", cell_value)
                    ):
                        name = cell_value.strip()
                        break

                if not name:
                    continue

                # Get user info once per person (optimization)
                user: Employee = await stp_repo.employee.get_users(fullname=name)
                if not user:
                    continue

                # Check duty information for all dates in the month
                for day, date_col in date_columns.items():
                    if date_col < len(df.columns):
                        duty_cell = self.utils.get_cell_value(df, row_idx, date_col)

                        if duty_cell and duty_cell.strip() not in [
                            "",
                            "nan",
                            "None",
                            "0",
                            "0.0",
                        ]:
                            shift_type, schedule = self.parse_duty_entry(duty_cell)

                            if shift_type in ["С", "П"] and self.utils.is_time_format(
                                schedule
                            ):
                                if day not in month_duties:
                                    month_duties[day] = []

                                month_duties[day].append(
                                    DutyInfo(
                                        name=name,
                                        user_id=user.user_id,
                                        username=user.username,
                                        schedule=schedule,
                                        shift_type=shift_type,
                                        work_hours=schedule,
                                    )
                                )

            total_duties = sum(len(duties) for duties in month_duties.values())
            logger.info(
                f"Found duties for {len(month_duties)} days in month {date.month}/{date.year}, total {total_duties} duties"
            )
            return month_duties

        except Exception as e:
            logger.debug(f"[Дежурные] Не удалось найти график дежурных: {e}")
            return {}

    async def get_duties_for_date(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> List[DutyInfo]:
        """Get list of duty officers for specified date."""
        try:
            # For НТП divisions, use separate seniority file
            if division in ["НТП", "НТП1", "НТП2"]:
                duty_file = self.file_manager.uploads_folder / "Старшинство_НТП.xlsx"
                if not duty_file.exists():
                    raise FileNotFoundError(
                        "Файл графика дежурных 'Старшинство_НТП.xlsx' не найден"
                    )
                schedule_file = duty_file
            # For НЦК division, use separate seniority file
            elif division == "НЦК":
                duty_file = self.file_manager.uploads_folder / "Старшинство_НЦК.xlsx"
                if not duty_file.exists():
                    raise FileNotFoundError(
                        "Файл графика дежурных 'Старшинство_НЦК.xlsx' не найден"
                    )
                schedule_file = duty_file
            else:
                schedule_file = self.file_manager.find_schedule_file(division)
                if not schedule_file:
                    raise FileNotFoundError(
                        f"Duty schedule file for {division} not found"
                    )

            sheet_name = self.get_duty_sheet_name(date)

            try:
                df = self.read_excel_file(schedule_file, sheet_name)
            except Exception as e:
                logger.warning(f"Failed to read schedule with primary sheet name: {e}")

                # For НТП and НЦК divisions, try different sheet name patterns
                if division in ["НТП", "НТП1", "НТП2", "НЦК"]:
                    # Try just the month name
                    month_names = [
                        "Январь",
                        "Февраль",
                        "Март",
                        "Апрель",
                        "Май",
                        "Июнь",
                        "Июль",
                        "Август",
                        "Сентябрь",
                        "Октябрь",
                        "Ноябрь",
                        "Декабрь",
                    ]
                    month_name = month_names[date.month - 1]
                    try:
                        df = self.read_excel_file(schedule_file, month_name)
                        logger.debug(
                            f"Successfully read {division} duty sheet with name: {month_name}"
                        )
                    except Exception as e2:
                        logger.warning(
                            f"Failed to read {division} duty sheet with month name '{month_name}': {e2}"
                        )
                        df = None
                else:
                    df = None

            if df is None:
                logger.debug(
                    f"[График дежурных] Не удалось найти график дежурных на {date} для {division}"
                )
                raise ValueError("Не удалось найти график дежурных")

            date_col = self.date_finder.find_date_column(df, date, search_rows=3)
            if date_col is None:
                logger.warning(f"Date {date.day} not found in duty schedule")
                return []

            duties = []

            for row_idx in range(len(df)):
                # Find name in first few columns
                name = ""
                for col_idx in range(min(3, len(df.columns))):
                    cell_value = self.utils.get_cell_value(df, row_idx, col_idx)

                    if (
                        len(cell_value.split()) >= 3
                        and re.search(r"[А-Яа-я]", cell_value)
                        and not re.search(r"\d", cell_value)
                    ):
                        name = cell_value.strip()
                        break

                if not name:
                    continue

                # Check duty information for this date
                if date_col < len(df.columns):
                    duty_cell = self.utils.get_cell_value(df, row_idx, date_col)

                    if duty_cell and duty_cell.strip() not in [
                        "",
                        "nan",
                        "None",
                        "0",
                        "0.0",
                    ]:
                        shift_type, schedule = self.parse_duty_entry(duty_cell)

                        if shift_type in ["С", "П"] and self.utils.is_time_format(
                            schedule
                        ):
                            user: Employee = await stp_repo.employee.get_users(
                                fullname=name
                            )
                            if user:
                                duties.append(
                                    DutyInfo(
                                        name=name,
                                        user_id=user.user_id,
                                        username=user.username,
                                        schedule=schedule,
                                        shift_type=shift_type,
                                        work_hours=schedule,
                                    )
                                )

            logger.info(
                f"Found {len(duties)} duty officers for {date.strftime('%d.%m.%Y')}"
            )
            return duties

        except Exception as e:
            logger.debug(f"[Дежурные] Не удалось найти график дежурных:  {e}")
            return []

    async def format_schedule(
        self,
        duties: List[DutyInfo],
        date: datetime,
        highlight_current: bool = False,
        division: str = None,
        stp_repo=None,
    ) -> str:
        """Format duties for date display."""
        if not duties:
            return f"<b>👮‍♂️ Дежурные • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдено дежурных на эту дату"

        lines = [f"<b>👮‍♂️ Дежурные • {date.strftime('%d.%m.%Y')}</b>\n"]

        # Get current duties if highlighting is needed
        current_senior = None
        current_helper = None
        if highlight_current and division and stp_repo:
            try:
                current_senior = await self.get_current_senior_duty(division, stp_repo)
                current_helper = await self.get_current_helper_duty(division, stp_repo)
            except Exception as e:
                logger.error(f"Error getting current duties: {e}")

        # Group by time
        time_groups = {}
        for duty in duties:
            time_schedule = duty.schedule
            if not time_schedule or not self.utils.is_time_format(time_schedule):
                continue

            if time_schedule not in time_groups:
                time_groups[time_schedule] = {"seniors": [], "helpers": []}

            if duty.shift_type == "С":
                time_groups[time_schedule]["seniors"].append(duty)
            elif duty.shift_type == "П":
                time_groups[time_schedule]["helpers"].append(duty)
            else:
                time_groups[time_schedule]["seniors"].append(duty)

        # Sort by time
        sorted_times = sorted(
            time_groups.keys(), key=lambda t: self.utils.parse_time_range(t)[0]
        )

        # Identify current duty time slots
        current_time_slots = set()
        if highlight_current and (current_senior or current_helper):
            for time_schedule in sorted_times:
                group = time_groups[time_schedule]
                for duty in group["seniors"] + group["helpers"]:
                    if (
                        current_senior
                        and duty.name == current_senior.name
                        and duty.shift_type == current_senior.shift_type
                    ) or (
                        current_helper
                        and duty.name == current_helper.name
                        and duty.shift_type == current_helper.shift_type
                    ):
                        current_time_slots.add(time_schedule)

        # Check if we need to start/end blockquote
        in_blockquote = False

        # Count current slots for spacing logic
        current_slots_count = len(current_time_slots)

        for i, time_schedule in enumerate(sorted_times):
            group = time_groups[time_schedule]
            is_current_slot = time_schedule in current_time_slots

            # Start blockquote if this is first current slot
            if is_current_slot and not in_blockquote:
                lines.append(f"<blockquote>⏰ {time_schedule}")
                in_blockquote = True
            # End blockquote if we were in one but this slot is not current
            elif not is_current_slot and in_blockquote:
                lines.append("</blockquote>")
                in_blockquote = False
                # Add time header without <b> tags
                lines.append(f"⏰ {time_schedule}")
            else:
                # Add time header without <b> tags
                lines.append(f"⏰ {time_schedule}")

            # Add senior officers
            for duty in group["seniors"]:
                lines.append(
                    f"Дежурный - {
                        format_fullname(
                            duty.name, True, True, duty.username, duty.user_id
                        )
                    }"
                )

            # Add helpers
            for duty in group["helpers"]:
                lines.append(
                    f"Помощник - {
                        format_fullname(
                            duty.name, True, True, duty.username, duty.user_id
                        )
                    }"
                )

            # Check if next slot is current to decide whether to close blockquote
            next_is_current = False
            if i + 1 < len(sorted_times):
                next_time_schedule = sorted_times[i + 1]
                next_is_current = next_time_schedule in current_time_slots

            # Add spacing logic
            if is_current_slot and not next_is_current and in_blockquote:
                # End blockquote if this was current slot and next is not current (or this is last slot)
                lines.append("</blockquote>")
                in_blockquote = False
            elif not is_current_slot:
                # Regular spacing for non-current slots
                lines.append("")
            elif is_current_slot and next_is_current:
                # Add empty line between current slots if there are few current slots (≤3)
                if current_slots_count <= 3:
                    lines.append("")

        # Remove last empty line
        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines)


class HeadScheduleParser(BaseExcelParser):
    """Parser for head schedules."""

    async def get_heads_for_date(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> List[HeadInfo]:
        """Get list of heads for specified date."""
        duty_parser = DutyScheduleParser()
        duties = await duty_parser.get_duties_for_date(date, division, stp_repo)

        try:
            # For НТП divisions, use НТП2 file since НТП1 doesn't contain head data
            if division in ["НТП", "НТП1", "НТП2"]:
                schedule_file = self.file_manager.find_schedule_file("НТП2")
            else:
                schedule_file = self.file_manager.find_schedule_file(division)

            if not schedule_file:
                raise FileNotFoundError(f"Head schedule file for {division} not found")

            df = self.read_excel_file(schedule_file, "ГРАФИК")
            if df is None:
                raise ValueError("Failed to read head schedule")

            date_col = self.date_finder.find_date_column(df, date)
            if date_col is None:
                logger.warning(f"Date {date.day} not found in head schedule")
                return []

            heads = []

            for row_idx in range(len(df)):
                position_found = False
                name = ""

                # Look for position and name in first few columns
                for col_idx in range(min(5, len(df.columns))):
                    cell_value = self.utils.get_cell_value(df, row_idx, col_idx)

                    if "Руководитель группы" in cell_value:
                        position_found = True

                    if (
                        not name
                        and len(cell_value.split()) >= 3
                        and re.search(r"[А-Яа-я]", cell_value)
                        and "Руководитель" not in cell_value
                    ):
                        name = cell_value.strip()

                if not position_found or not name:
                    continue

                # Check schedule for this date
                if date_col < len(df.columns):
                    schedule_cell = self.utils.get_cell_value(df, row_idx, date_col)
                    if schedule_cell and schedule_cell.strip():
                        if self.utils.is_time_format(schedule_cell):
                            duty_info = await self._check_duty_for_head(name, duties)
                            user: Employee = await stp_repo.employee.get_users(
                                fullname=name
                            )
                            if user:
                                heads.append(
                                    HeadInfo(
                                        name=name,
                                        user_id=user.user_id,
                                        username=user.username,
                                        schedule=schedule_cell.strip(),
                                        duty_info=duty_info,
                                    )
                                )

            logger.info(f"Found {len(heads)} heads for {date.strftime('%d.%m.%Y')}")
            return heads

        except Exception as e:
            logger.error(f"Error getting heads: {e}")
            return []

    async def _check_duty_for_head(
        self, head_name: str, duties: List[DutyInfo]
    ) -> Optional[str]:
        """Check if head is on duty."""
        try:
            for duty in duties:
                if self.utils.names_match(head_name, duty.name):
                    return f"{duty.schedule} {duty.shift_type}"
            return None
        except Exception as e:
            logger.debug(f"Error checking duty for {head_name}: {e}")
            return None

    def format_schedule(self, heads: List[HeadInfo], date: datetime) -> str:
        """Format heads schedule for display."""
        if not heads:
            return f"<b>👑 Руководители • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдено руководителей на эту дату"

        lines = [f"<b>👑 Руководители • {date.strftime('%d.%m.%Y')}</b>\n"]

        # Group by time
        time_groups = {}
        for head in heads:
            time_schedule = head.schedule
            if not time_schedule or not self.utils.is_time_format(time_schedule):
                continue

            time_match = re.search(r"(\d{1,2}:\d{2}-\d{1,2}:\d{2})", time_schedule)
            time_key = time_match.group(1) if time_match else time_schedule

            if time_key not in time_groups:
                time_groups[time_key] = []
            time_groups[time_key].append(head)

        # Sort by time
        sorted_times = sorted(
            time_groups.keys(), key=lambda t: self.utils.parse_time_range(t)[0]
        )

        for time_schedule in sorted_times:
            group_heads = time_groups[time_schedule]
            lines.append(f"⏰ <b>{time_schedule}</b>")

            for head in group_heads:
                head_line = f"{
                    format_fullname(head.name, True, True, head.username, head.user_id)
                }"

                if head.duty_info:
                    head_line += f" ({head.duty_info})"

                lines.append(head_line)

            lines.append("")

        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines)


class GroupScheduleParser(BaseExcelParser):
    """Parser for group schedules."""

    def __init__(self, uploads_folder: str = "uploads"):
        super().__init__(uploads_folder)
        self.duty_parser = DutyScheduleParser(uploads_folder)

    def _group_members_by_start_time(
        self, members: List[GroupMemberInfo]
    ) -> Dict[str, List[GroupMemberInfo]]:
        """Group members by their start time."""
        grouped = defaultdict(list)

        for member in members:
            start_time = self.time_utils.extract_start_time(member.working_hours)
            grouped[start_time].append(member)

        return dict(grouped)

    def _format_member_with_link(self, member: GroupMemberInfo) -> str:
        """Format member name with link and working hours."""
        user_link = format_fullname(
            member.name, True, True, member.username, member.user_id
        )

        # Add working hours
        working_hours = member.working_hours or "Не указано"
        result = f"{user_link} <code>{working_hours}</code>"

        # Add duty information if available
        if member.duty_info:
            result += f" ({member.duty_info})"

        return result

    def _sort_members_by_time(
        self, members: List[GroupMemberInfo]
    ) -> List[GroupMemberInfo]:
        """Sort members by start time."""
        return sorted(
            members,
            key=lambda m: self.time_utils.parse_time_for_sorting(
                self.time_utils.extract_start_time(m.working_hours)
            ),
        )

    def format_schedule(self, members: List[GroupMemberInfo], date: datetime) -> str:
        """Format group schedule for display."""
        if not members:
            return f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдены участники группы"

        # Group by start time
        grouped_by_start_time = self._group_members_by_start_time(members)

        # Sort groups by time
        sorted_start_times = sorted(
            grouped_by_start_time.keys(), key=self.time_utils.parse_time_for_sorting
        )

        lines = [f"👥 <b>Group for {date.strftime('%d.%m.%Y')}</b>", ""]

        for start_time in sorted_start_times:
            members_group = grouped_by_start_time[start_time]
            lines.append(f"🕒 <b>{start_time}</b>")

            for member in members_group:
                lines.append(self._format_member_with_link(member))

            lines.append("")

        # Remove last empty line
        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines)

    async def get_group_members_for_head(
        self, head_fullname: str, date: datetime, division: str, stp_repo
    ) -> List[GroupMemberInfo]:
        """Get list of group members for a head."""
        try:
            group_members = []

            # For НТП divisions, process both НТП1 and НТП2 files
            if "НТП" in division:
                divisions_to_check = ["НТП1", "НТП2"]
            else:
                divisions_to_check = [division]

            for div in divisions_to_check:
                schedule_file = self.file_manager.find_schedule_file(div)
                if not schedule_file:
                    logger.warning(f"Schedule file for {div} not found")
                    continue

                df = self.read_excel_file(schedule_file)
                if df is None:
                    logger.warning(f"Failed to read schedule file for {div}")
                    continue

                header_info = self._get_header_columns()
                date_column = self.date_finder.find_date_column(df, date)

                # Process members from this division file
                division_members = await self._process_division_members(
                    df, head_fullname, date_column, header_info, stp_repo
                )
                group_members.extend(division_members)

            # Fetch duty information for all members
            if group_members:
                for div in divisions_to_check:
                    try:
                        duties = await self.duty_parser.get_duties_for_date(
                            date, div, stp_repo
                        )
                        for member in group_members:
                            # Check if this member is on duty (skip if already has duty info)
                            if not hasattr(member, "duty_info") or not member.duty_info:
                                for duty in duties:
                                    if self.utils.names_match(member.name, duty.name):
                                        member.duty_info = (
                                            f"{duty.schedule} {duty.shift_type}"
                                        )
                                        break
                    except Exception as duty_error:
                        logger.warning(
                            f"Could not fetch duty information for {div}: {duty_error}"
                        )

            logger.info(
                f"Found {len(group_members)} total members for head {head_fullname} across divisions"
            )
            return self._sort_members_by_time(group_members)

        except Exception as e:
            logger.error(f"Error getting group members for {head_fullname}: {e}")
            return []

    async def _process_division_members(
        self, df, head_fullname: str, date_column, header_info, stp_repo
    ) -> List[GroupMemberInfo]:
        """Process members from a single division file."""
        division_members = []

        for row_idx in range(header_info["header_row"] + 1, len(df)):
            name_cell = self.utils.get_cell_value(df, row_idx, 0)
            schedule_cell = self.utils.get_cell_value(
                df, row_idx, header_info.get("schedule_col", 1)
            )
            position_cell = self.utils.get_cell_value(
                df, row_idx, header_info.get("position_col", 4)
            )
            head_cell = self.utils.get_cell_value(
                df, row_idx, header_info.get("head_col", 5)
            )

            # Check if this person belongs to the specified head
            if not self.utils.names_match(head_fullname, head_cell):
                continue

            if not self.utils.is_valid_name(name_cell):
                continue

            # Get working hours for the specific date
            working_hours = "Не указано"
            if date_column is not None:
                hours_cell = self.utils.get_cell_value(df, row_idx, date_column)
                if hours_cell and hours_cell.strip():
                    if self.utils.is_time_format(hours_cell):
                        working_hours = hours_cell
                    else:
                        # Non-time value (vacation, day off, etc.) - skip this person
                        continue
                else:
                    # Empty cell means day off - skip this person
                    continue

            # Get user from database
            user = None
            try:
                user = await stp_repo.employee.get_users(fullname=name_cell.strip())
            except Exception as e:
                logger.debug(f"Error getting user {name_cell}: {e}")

            if not user:
                logger.debug(f"User {name_cell.strip()} not found in DB, skipping")
                continue

            member = GroupMemberInfo(
                name=name_cell.strip(),
                user_id=user.user_id,
                username=user.username,
                schedule=schedule_cell.strip() if schedule_cell else "Не указано",
                position=position_cell.strip() if position_cell else "Специалист",
                working_hours=working_hours,
            )

            division_members.append(member)

        return division_members

    async def get_group_members_for_user(
        self, user_fullname: str, date: datetime, division: str, stp_repo
    ) -> List[GroupMemberInfo]:
        """Get list of group colleagues for a regular user."""
        try:
            user = await stp_repo.employee.get_users(fullname=user_fullname)
            if not user or not user.head:
                logger.warning(
                    f"User {user_fullname} not found or has no head assigned"
                )
                return []

            # Get all members under the same head
            all_members = await self.get_group_members_for_head(
                user.head, date, division, stp_repo
            )

            return self._sort_members_by_time(all_members)

        except Exception as e:
            logger.error(f"Error getting colleagues for {user_fullname}: {e}")
            return []

    @staticmethod
    def _get_header_columns() -> dict:
        """Find header columns in the DataFrame."""
        # This is a simplified implementation - adjust based on actual Excel structure
        return {
            "header_row": 0,
            "schedule_col": 1,
            "position_col": 4,
            "head_col": 5,
        }

    def format_group_schedule_for_head(
        self,
        date: datetime,
        group_members: List[GroupMemberInfo],
        page: int = 1,
        members_per_page: int = 20,
    ) -> tuple[str, int, bool, bool]:
        """Format group schedule for head with pagination."""
        if not group_members:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдены участники группы",
                1,
                False,
                False,
            )

        # Apply pagination
        total_members = len(group_members)
        total_pages = max(1, (total_members + members_per_page - 1) // members_per_page)

        start_idx = (page - 1) * members_per_page
        end_idx = start_idx + members_per_page
        page_members = group_members[start_idx:end_idx]

        # Group by start time
        grouped_by_start_time = self._group_members_by_start_time(page_members)
        sorted_start_times = sorted(
            grouped_by_start_time.keys(), key=self.time_utils.parse_time_for_sorting
        )

        # Build message
        lines = [f"❤️ <b>Твоя группа • {date.strftime('%d.%m.%Y')}</b>"]

        # Add pagination info
        if total_pages > 1:
            lines.append(
                f"<i>Страница {page} из {total_pages}\nОтображено {len(page_members)} из {total_members} участников</i>"
            )

        lines.append("")

        for start_time in sorted_start_times:
            members = grouped_by_start_time[start_time]
            lines.append(f"🕒 <b>{start_time}</b>")

            for member in members:
                lines.append(self._format_member_with_link(member))

            lines.append("")

        # Remove last empty line
        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines), total_pages, page > 1, page < total_pages

    def format_group_schedule_for_user(
        self,
        date: datetime,
        group_members: List[GroupMemberInfo],
        page: int = 1,
        members_per_page: int = 20,
    ) -> tuple[str, int, bool, bool]:
        """Format group schedule for regular user with pagination."""
        if not group_members:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ График для группы не найден",
                1,
                False,
                False,
            )

        colleagues = [member for member in group_members]

        if not colleagues:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ График для группы не найден",
                1,
                False,
                False,
            )

        # Apply pagination
        total_colleagues = len(group_members)
        total_pages = max(
            1, (total_colleagues + members_per_page - 1) // members_per_page
        )

        start_idx = (page - 1) * members_per_page
        end_idx = start_idx + members_per_page
        page_colleagues = colleagues[start_idx:end_idx]

        # Group by start time
        grouped_by_start_time = self._group_members_by_start_time(page_colleagues)
        sorted_start_times = sorted(
            grouped_by_start_time.keys(), key=self.time_utils.parse_time_for_sorting
        )

        # Build message
        lines = [f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>"]

        # Add pagination info
        if total_pages > 1:
            lines.append(
                f"<i>Страница {page} из {total_pages}\nОтображено {len(page_colleagues)} из {total_colleagues} участников</i>"
            )

        lines.append("")

        for start_time in sorted_start_times:
            members = grouped_by_start_time[start_time]
            lines.append(f"<b>{start_time}</b>")

            for member in members:
                lines.append(self._format_member_with_link(member))

            lines.append("")

        # Remove last empty line
        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines), total_pages, page > 1, page < total_pages


# Factory class for creating parsers
class ScheduleParserFactory:
    """Factory for creating schedule parsers."""

    @staticmethod
    def create_schedule_parser(uploads_folder: str = "uploads") -> ScheduleParser:
        """Create a schedule parser instance."""
        return ScheduleParser(uploads_folder)

    @staticmethod
    def create_duty_parser(uploads_folder: str = "uploads") -> DutyScheduleParser:
        """Create a duty schedule parser instance."""
        return DutyScheduleParser(uploads_folder)

    @staticmethod
    def create_head_parser(uploads_folder: str = "uploads") -> HeadScheduleParser:
        """Create a head schedule parser instance."""
        return HeadScheduleParser(uploads_folder)

    @staticmethod
    def create_group_parser(uploads_folder: str = "uploads") -> GroupScheduleParser:
        """Create a group schedule parser instance."""
        return GroupScheduleParser(uploads_folder)

    @staticmethod
    async def get_current_senior_duty(
        division: str, stp_repo, uploads_folder: str = "uploads"
    ) -> Optional[DutyInfo]:
        """Get current senior duty for division - convenience method."""
        parser = ScheduleParserFactory.create_duty_parser(uploads_folder)
        return await parser.get_current_senior_duty(division, stp_repo)
