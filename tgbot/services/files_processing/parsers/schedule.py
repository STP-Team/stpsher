"""Модуль парсеров графиков.

Парсеры для работы с:
- Графиками сотрудников (ScheduleParser)
- Графиками дежурных (DutyScheduleParser)
- Графиками руководителей (HeadScheduleParser)
- Групповыми графиками (GroupScheduleParser)
"""

import calendar
import logging
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from aiogram import Bot
from aiogram.utils.deep_linking import create_start_link
from openpyxl import load_workbook
from stp_database import Employee, MainRequestsRepo

from tgbot.misc.dicts import schedule_types
from tgbot.misc.helpers import format_fullname, tz_perm

from ..core.analyzers import ScheduleAnalyzer
from ..core.excel import ExcelReader
from ..core.models import DutyInfo, GroupMemberInfo, HeadInfo
from ..formatters.schedule import ScheduleFormatter
from ..managers.files import MonthManager
from .base import BaseParser

logger = logging.getLogger(__name__)


class ScheduleParser(BaseParser):
    """Парсер графиков."""

    def __init__(self, uploads_folder: str = "uploads"):
        super().__init__(uploads_folder)
        self.analyzer = ScheduleAnalyzer()
        self.formatter = ScheduleFormatter()

    def get_user_schedule(
        self, fullname: str, month: str, division: str
    ) -> Dict[str, str]:
        """Получает график пользователя.

        Args:
            fullname: ФИО пользователя
            month: Название месяца
            division: Направление

        Returns:
            Словарь с графиком {день: значение}

        Raises:
            FileNotFoundError: Если файл графика не найден
        """
        try:
            schedule_file = self.file_manager.find_schedule_file(division)
            if not schedule_file:
                raise FileNotFoundError(f"Файл графика для {division} не найден")

            reader = ExcelReader(schedule_file)
            schedule = reader.extract_user_schedule(fullname, month)

            logger.info(
                f"[Excel] Найдено {len(schedule)} дней для {fullname} в {month}"
            )
            return schedule

        except Exception as e:
            logger.error(f"[Excel] Ошибка нахождения графика: {e}")
            raise

    async def get_user_schedule_with_duties(
        self,
        fullname: str,
        month: str,
        division: str,
        stp_repo=None,
        current_day_only: bool = False,
        bot: Bot = None,
    ) -> Dict[str, Tuple[str, Optional[str]]]:
        """Получает график пользователя с информацией о дежурствах и обменах.

        Args:
            fullname: ФИО пользователя
            month: Название месяца
            division: Направление
            stp_repo: Репозиторий операций с базой STP
            current_day_only: Если True, получает дежурство только для текущего дня
            bot: Экземпляр бота

        Returns:
            Словарь с маппингом день -> (график, информация_о_дежурстве_и_обменах)
        """
        try:
            schedule_data = self.get_user_schedule(fullname, month, division)

            if not schedule_data or not stp_repo:
                return {
                    day: (schedule, None) for day, schedule in schedule_data.items()
                }

            # Получаем парсер графика дежурных
            duty_parser = DutyScheduleParser()

            # Получаем данные о текущем дне
            current_year = datetime.now().year
            current_month_num = datetime.now().month
            current_day_num = datetime.now().day
            month_num = MonthManager.get_month_number(month)

            # Проверяем открыт ли текущий месяц
            is_current_month = (current_year, month_num) == (
                current_year,
                current_month_num,
            )

            month_duties = {}

            # Для компактного режима грузим дежурных только за текущий день
            if current_day_only and is_current_month:
                try:
                    current_date = datetime(current_year, month_num, current_day_num)
                    today_duties = await duty_parser.get_duties_for_date(
                        current_date, division, stp_repo
                    )
                    if today_duties:
                        month_duties[current_day_num] = today_duties
                    logger.debug(
                        f"[Excel] Получены дежурные только на текущий день ({current_day_num})"
                    )
                except Exception as e:
                    logger.debug(
                        f"[Excel] Ошибка получения дежурных на текущий день: {e}"
                    )
            else:
                # Дежурные на весь месяц (для детального просмотра или не текущего месяца)
                try:
                    first_day_of_month = datetime(current_year, month_num, 1)
                    month_duties = await duty_parser.get_duties_for_month(
                        first_day_of_month, division, stp_repo
                    )
                    logger.debug(
                        f"[Excel] Получены дежурные на {len(month_duties)} дней на {month}"
                    )
                except Exception as e:
                    logger.warning(f"[Excel] Ошибка получения дежурных на месяц: {e}")
                    month_duties = {}

            # Получаем информацию о купленных обменах пользователя
            schedule_exchanges = {}
            if stp_repo:
                try:
                    user = await stp_repo.employee.get_users(fullname=fullname)
                    if user:
                        # Получаем купленные обмены пользователя (где он buyer)
                        user_exchanges = await stp_repo.exchange.get_user_exchanges(
                            user_id=user.user_id,
                            status="sold",
                        )

                        current_year = datetime.now().year
                        month_num = MonthManager.get_month_number(month)

                        for exchange in user_exchanges:
                            if (
                                exchange.owner_id == user.user_id
                                and exchange.in_owner_schedule
                            ) or (
                                exchange.counterpart_id == user.user_id
                                and exchange.in_counterpart_schedule
                            ):
                                # Проверяем что обмен относится к нужному месяцу
                                if (
                                    exchange.start_time.year == current_year
                                    and exchange.start_time.month == month_num
                                ):
                                    day_num = exchange.start_time.day
                                    emoji = (
                                        "📈"
                                        if exchange.owner_id == user.user_id
                                        and exchange.owner_intent == "buy"
                                        else "📉"
                                    )
                                    if bot:
                                        deeplink = await create_start_link(
                                            bot=bot,
                                            payload=f"exchange_{exchange.id}",
                                            encode=True,
                                        )
                                        exchange_info = f"<a href='{deeplink}'>{emoji} {exchange.start_time.strftime('%H:%M')}-{exchange.end_time.strftime('%H:%M')}</a>"
                                    else:
                                        exchange_info = f"{emoji} {exchange.start_time.strftime('%H:%M')}-{exchange.end_time.strftime('%H:%M')}"
                                    schedule_exchanges[day_num] = exchange_info

                        logger.debug(
                            f"[Excel] Найдено {len(user_exchanges)} сделок для {fullname}"
                        )

                except Exception as e:
                    logger.debug(
                        f"[Excel] Ошибка получения обменов для {fullname}: {e}"
                    )

            schedule_with_duties = {}

            for day, schedule in schedule_data.items():
                duty_info = None
                exchange_info = None

                try:
                    day_match = re.search(r"(\d+)", day)
                    if day_match:
                        day_num = int(day_match.group(1))

                        # Проверяем дежурства
                        if month_duties and day_num in month_duties:
                            duties = month_duties[day_num]

                            # Проверяем есть ли сотрудник в дежурных
                            for duty in duties:
                                if self.names_match(fullname, duty.name):
                                    duty_info = f"{duty.schedule} {duty.shift_type}"
                                    break

                        # Проверяем сделки
                        if day_num in schedule_exchanges:
                            exchange_info = schedule_exchanges[day_num]

                except Exception as e:
                    logger.debug(
                        f"[Excel] Ошибка проверки дежурного/обмена для {fullname} на {day}: {e}"
                    )

                # Объединяем информацию о дежурствах и обменах
                combined_info = None
                if duty_info and exchange_info:
                    combined_info = f"{duty_info} | {exchange_info}"
                elif duty_info:
                    combined_info = duty_info
                elif exchange_info:
                    combined_info = exchange_info

                schedule_with_duties[day] = (schedule, combined_info)

            return schedule_with_duties

        except Exception as e:
            logger.error(f"[Excel] Ошибка получения графика с дежурными: {e}")
            schedule_data = self.get_user_schedule(fullname, month, division)
            return {day: (schedule, None) for day, schedule in schedule_data.items()}

    def get_user_schedule_formatted(
        self,
        fullname: str,
        month: str,
        division: str,
        compact: bool = False,
    ) -> str:
        """Получает отформатированный график пользователя.

        Args:
            fullname: ФИО пользователя
            month: Название месяца
            division: Направление
            compact: Использовать компактный формат

        Returns:
            Отформатированная строка с графиком
        """
        try:
            schedule_data = self.get_user_schedule(fullname, month, division)

            if not schedule_data:
                return f"❌ Не найден график для <b>{fullname}</b> на {month}"

            analysis = self.analyzer.analyze_schedule(schedule_data)

            if compact:
                return self.formatter.format_compact(month, *analysis)
            else:
                return self.formatter.format_detailed(month, *analysis)

        except Exception as e:
            logger.error(f"[Excel] Ошибка форматирования графика: {e}")
            return f"❌ <b>Ошибка графика:</b>\n<code>{e}</code>"

    async def get_user_schedule_formatted_with_duties(
        self,
        fullname: str,
        month: str,
        division: str,
        compact: bool = False,
        stp_repo=None,
        bot: Bot = None,
    ) -> str:
        """Получает отформатированный график пользователя с дежурствами и обменами.

        Для компактного вида: получает дежурство только текущего дня (быстро)
        Для детального вида: получает дежурства за весь месяц (медленнее, но полно)

        Args:
            fullname: ФИО пользователя
            month: Название месяца
            division: Направление
            compact: Использовать компактный формат
            stp_repo: Репозиторий операций с базой STP
            bot: Экземпляр бота

        Returns:
            Отформатированная строка с графиком, дежурствами и обменами
        """
        try:
            schedule_data_with_duties = await self.get_user_schedule_with_duties(
                fullname,
                month,
                division,
                stp_repo,
                current_day_only=compact,
                bot=bot,
            )

            if not schedule_data_with_duties:
                return f"❌ Не найден график для <b>{fullname}</b> на {month}"

            schedule_data = {
                day: schedule
                for day, (schedule, _) in schedule_data_with_duties.items()
            }
            analysis = self.analyzer.analyze_schedule(schedule_data)

            if compact:
                # Extract current day's duty and exchange info for compact view
                current_day_info = None
                current_day_num = datetime.now().day
                for day_key, (_, combined_info) in schedule_data_with_duties.items():
                    day_match = re.search(r"(\d+)", day_key)
                    if day_match and int(day_match.group(1)) == current_day_num:
                        current_day_info = combined_info
                        break

                logger.debug(
                    f"[Excel] Компактный вид дежурных/обменов: {current_day_info or 'None'}"
                )

                # Для компактного вида отображаем только текущий день дежурств и обменов
                return self.formatter.format_compact(
                    month, *analysis, current_day_duty=current_day_info
                )
            else:
                # Для детального вида отображаем график дежурств на весь месяц
                return self.formatter.format_detailed_with_duties(
                    month, schedule_data_with_duties, *analysis
                )

        except Exception as e:
            logger.error(f"Schedule formatting error: {e}")
            return f"❌ <b>Ошибка графика:</b>\n<code>{e}</code>"

    @staticmethod
    def _is_additional_shift_color(cell) -> bool:
        """Проверяет, является ли цвет ячейки цветом дополнительной смены (#cc99ff).

        Args:
            cell: Ячейка openpyxl

        Returns:
            True если ячейка имеет цвет дополнительной смены
        """
        try:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color:
                    # Check for #cc99ff color (with or without alpha channel)
                    color_str = str(color).upper()
                    return "CC99FF" in color_str
        except Exception:
            pass
        return False

    def get_user_schedule_with_additional_shifts(
        self, fullname: str, month: str, division: str
    ) -> Tuple[Dict[str, str], Dict[str, str]]:
        """Получает график пользователя с разделением на обычные и дополнительные смены.

        Дополнительные смены определяются по цвету ячейки #cc99ff.

        Args:
            fullname: ФИО пользователя
            month: Название месяца
            division: Направление

        Returns:
            Кортеж (обычный_график, дополнительные_смены)

        Raises:
            FileNotFoundError: Если файл графика не найден
            ValueError: Если пользователь не найден
        """
        try:
            schedule_file = self.file_manager.find_schedule_file(division)
            if not schedule_file:
                raise FileNotFoundError(f"Файл графика для {division} не найден")

            # Load with openpyxl to access cell formatting
            wb = load_workbook(schedule_file, data_only=False)
            ws = wb["ГРАФИК"] if "ГРАФИК" in wb.sheetnames else wb.active

            # Use ExcelReader for data access
            reader = ExcelReader(schedule_file)

            # Find user row
            user_row = reader.find_user_row(fullname)
            if user_row is None:
                raise ValueError(f"Сотрудник {fullname} не найден в графике")

            # Get month range
            month_range = reader.get_month_range(month)
            if month_range is None:
                raise ValueError(f"Месяц {month} не найден")

            start_col, end_col = month_range
            day_headers = reader.get_day_headers(start_col, end_col)

            schedule = {}
            additional_shifts = {}

            for col_idx, day in day_headers.items():
                cell_value = reader.get_cell(user_row, col_idx)
                schedule_value = cell_value.strip() if cell_value is not None else ""

                if schedule_value.lower() in schedule_types["day_off"]:
                    schedule_value = None

                # Check cell color in openpyxl (1-indexed)
                cell = ws.cell(row=user_row + 1, column=col_idx + 1)
                is_additional_shift = self._is_additional_shift_color(cell)

                if (
                    is_additional_shift
                    and schedule_value not in schedule_types["day_off"]
                ):
                    additional_shifts[day] = schedule_value
                else:
                    schedule[day] = schedule_value

            logger.info(
                f"[Optimized] Found {len(schedule)} regular days and {len(additional_shifts)} additional shifts for {fullname} in {month}"
            )
            return schedule, additional_shifts

        except Exception as e:
            logger.error(f"Error getting schedule with additional shifts: {e}")
            raise

    def parse(self, *args, **kwargs):
        """Реализация абстрактного метода parse.

        Returns:
            Результат get_user_schedule
        """
        return self.get_user_schedule(*args, **kwargs)


class DutyScheduleParser(BaseParser):
    """Полностью оптимизированный парсер графика дежурных с кэшированием."""

    @staticmethod
    def get_duty_sheet_name(date: datetime) -> str:
        """Генерирует название листа для графика дежурных.

        Args:
            date: Дата для определения месяца

        Returns:
            Название листа (например, "Дежурство Январь")
        """
        month_names = [
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        ]
        month_name = month_names[date.month - 1]
        return f"Дежурство {month_name}"

    @staticmethod
    def parse_duty_entry(cell_value: str) -> Tuple[str, str]:
        """Разбирает запись дежурства для извлечения типа смены и времени.

        Args:
            cell_value: Значение ячейки (например, "П 09:00-18:00")

        Returns:
            Кортеж (тип_смены, график)
        """
        if not cell_value or cell_value.strip() in ["", "nan", "None"]:
            return "", ""

        cell_value = cell_value.strip()

        if cell_value.startswith("П "):
            return "П", cell_value[2:].strip()
        elif cell_value.startswith("С "):
            return "С", cell_value[2:].strip()
        else:
            if re.search(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}", cell_value):
                return "", cell_value
            else:
                return "", cell_value

    async def get_duties_for_month(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> Dict[int, List[DutyInfo]]:
        """Получает все дежурства за весь месяц (оптимизировано с кэшированием).

        Args:
            date: Дата в нужном месяце
            division: Направление
            stp_repo: Репозиторий операций с базой STP

        Returns:
            Словарь {день_месяца: список_дежурных}
        """
        try:
            # Определяем файл дежурств
            if division in ["НТП1", "НТП2"]:
                duty_file = self.file_manager.uploads_folder / "Старшинство_НТП.xlsx"
            elif division == "НЦК":
                duty_file = self.file_manager.uploads_folder / "Старшинство_НЦК.xlsx"
            else:
                duty_file = self.file_manager.find_schedule_file(division)

            if not duty_file or not duty_file.exists():
                raise FileNotFoundError(
                    f"Файл график для дежурных {division} не найден"
                )

            sheet_name = self.get_duty_sheet_name(date)

            # Читаем файл с графиком
            reader = ExcelReader(duty_file, sheet_name)
            df = reader.df
            month_duties = {}

            # Находим все колонки с датами для месяца
            days_in_month = calendar.monthrange(date.year, date.month)[1]

            # Собираем все уникальные ФИО
            names_to_fetch = set()
            row_name_map = {}  # Map row_idx to name

            for row_idx in range(len(df)):
                # Находим ФИО сотрудника в первых колонках
                name = ""
                for col_idx in range(min(3, df.shape[1])):
                    cell_value = reader.get_cell(row_idx, col_idx)

                    if (
                        len(cell_value.split()) >= 3
                        and re.search(r"[А-Яа-я]", cell_value)
                        and not re.search(r"\d", cell_value)
                    ):
                        name = cell_value.strip()
                        break

                if name:
                    names_to_fetch.add(name)
                    row_name_map[row_idx] = name

            if not names_to_fetch:
                return {}

            employee_cache = {}
            for name in names_to_fetch:
                try:
                    user: Employee = await stp_repo.employee.get_users(fullname=name)
                    if user:
                        employee_cache[name] = user
                except Exception as e:
                    logger.debug(f"Error fetching user {name}: {e}")

            if not employee_cache:
                return {}

            date_column_cache = {}
            for day in range(1, days_in_month + 1):
                day_date = datetime(date.year, date.month, day)
                day_col = reader.find_date_column(day_date)
                if day_col is not None:
                    date_column_cache[day] = day_col

            # Сканим все строки и колонки
            for row_idx, name in row_name_map.items():
                # Пропускаем если сотрудник не найден в кеше
                if name not in employee_cache:
                    continue

                user = employee_cache[name]

                # Проверяем все дни в месяце
                for day, day_col in date_column_cache.items():
                    try:
                        if day_col < df.shape[1]:
                            duty_cell = reader.get_cell(row_idx, day_col)

                            if duty_cell and duty_cell.strip() not in [
                                "",
                                "nan",
                                "None",
                                "0",
                                "0.0",
                            ]:
                                shift_type, schedule = self.parse_duty_entry(duty_cell)

                                if shift_type in ["С", "П"] and self.is_time_format(
                                    schedule
                                ):
                                    if day not in month_duties:
                                        month_duties[day] = []

                                    month_duties[day].append(
                                        DutyInfo(
                                            name=name,
                                            user_id=user.user_id,
                                            username=user.username,
                                            schedule=schedule,
                                            shift_type=shift_type,
                                            work_hours=schedule,
                                        )
                                    )
                    except (ValueError, Exception):
                        continue

            total_duties = sum(len(duties) for duties in month_duties.values())
            logger.debug(
                f"[Excel] Найдены дежурные на {len(month_duties)} дней, всего {total_duties} дежурных"
            )
            return month_duties

        except Exception as e:
            logger.debug(f"[Excel] Ошибка проверки дежурных на месяц: {e}")
            return {}

    async def get_duties_for_date(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> List[DutyInfo]:
        """Получает дежурства для конкретной даты.

        Args:
            date: Дата
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Список дежурных на эту дату
        """
        try:
            # Сперва пробуем использовать кеш
            month_duties = await self.get_duties_for_month(date, division, stp_repo)

            if date.day in month_duties:
                return month_duties[date.day]

            return []

        except Exception as e:
            logger.debug(f"[Excel] Ошибка получения дежурных для даты: {e}")
            return []

    async def get_current_senior_duty(
        self, division: str, stp_repo: MainRequestsRepo
    ) -> Optional[DutyInfo]:
        """Получает текущего старшего дежурного.

        Args:
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Информация о текущем старшем дежурном или None
        """
        date = datetime.now(tz_perm)

        try:
            duties = await self.get_duties_for_date(date, division, stp_repo)
            senior_duties = [duty for duty in duties if duty.shift_type == "С"]

            if not senior_duties:
                return None

            # Находим текущего дежурного по времени
            for duty in senior_duties:
                if self.is_time_format(duty.schedule):
                    start_minutes, end_minutes = self.parse_time_range(duty.schedule)
                    current_time_minutes = date.hour * 60 + date.minute

                    if start_minutes <= current_time_minutes <= end_minutes:
                        return duty

                    # Обрабатываем ночные смены
                    if end_minutes > 24 * 60:
                        if (
                            current_time_minutes >= start_minutes
                            or current_time_minutes <= (end_minutes - 24 * 60)
                        ):
                            return duty

            return None

        except Exception as e:
            logger.error(f"[Excel] Ошибка получения текущего дежурного: {e}")
            return None

    async def get_current_helper_duty(
        self, division: str, stp_repo: MainRequestsRepo
    ) -> Optional[DutyInfo]:
        """Получает текущего помощника дежурного.

        Args:
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Информация о текущем помощнике дежурного или None
        """
        date = datetime.now(tz_perm)

        try:
            duties = await self.get_duties_for_date(date, division, stp_repo)
            helper_duties = [duty for duty in duties if duty.shift_type == "П"]

            if not helper_duties:
                return None

            # Находим текущего помощника по времени
            for duty in helper_duties:
                if self.is_time_format(duty.schedule):
                    start_minutes, end_minutes = self.parse_time_range(duty.schedule)
                    current_time_minutes = date.hour * 60 + date.minute

                    if start_minutes <= current_time_minutes <= end_minutes:
                        return duty

                    # Обрабатываем ночные смены
                    if end_minutes > 24 * 60:
                        if (
                            current_time_minutes >= start_minutes
                            or current_time_minutes <= (end_minutes - 24 * 60)
                        ):
                            return duty

            return None

        except Exception as e:
            logger.error(f"[Excel] Ошибка получения текущего помощника: {e}")
            return None

    async def format_schedule(
        self,
        duties: List[DutyInfo],
        date: datetime,
        highlight_current: bool = False,
        division: str = None,
        stp_repo=None,
    ) -> str:
        """Форматирует дежурства для отображения.

        Args:
            duties: Список дежурных
            date: Дата
            highlight_current: Подсвечивать текущих дежурных
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Отформатированная строка с дежурствами
        """
        if not duties:
            return f"<b>👮‍♂️ Дежурные • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдено дежурных на эту дату"

        lines = [f"<b>👮‍♂️ Дежурные • {date.strftime('%d.%m.%Y')}</b>\n"]

        current_senior = None
        current_helper = None
        if highlight_current:
            current_senior = await self.get_current_senior_duty(
                division=division, stp_repo=stp_repo
            )
            current_helper = await self.get_current_helper_duty(
                division=division, stp_repo=stp_repo
            )

        # Группируем по времени
        time_groups = {}
        for duty in duties:
            time_schedule = duty.schedule
            if not time_schedule or not self.is_time_format(time_schedule):
                continue

            if time_schedule not in time_groups:
                time_groups[time_schedule] = {"seniors": [], "helpers": []}

            if duty.shift_type == "С":
                time_groups[time_schedule]["seniors"].append(duty)
            elif duty.shift_type == "П":
                time_groups[time_schedule]["helpers"].append(duty)
            else:
                time_groups[time_schedule]["seniors"].append(duty)

        # Сортировка по времени
        sorted_times = sorted(
            time_groups.keys(), key=lambda t: self.parse_time_range(t)[0]
        )

        # Идентификация тайм слотов текущих дежурных
        current_time_slots = set()
        if highlight_current and (current_senior or current_helper):
            for time_schedule in sorted_times:
                group = time_groups[time_schedule]
                for duty in group["seniors"] + group["helpers"]:
                    if (
                        current_senior
                        and duty.name == current_senior.name
                        and duty.shift_type == current_senior.shift_type
                    ) or (
                        current_helper
                        and duty.name == current_helper.name
                        and duty.shift_type == current_helper.shift_type
                    ):
                        current_time_slots.add(time_schedule)

        # Форматируем вывод
        in_blockquote = False
        current_slots_count = len(current_time_slots)

        for i, time_schedule in enumerate(sorted_times):
            group = time_groups[time_schedule]
            is_current_slot = time_schedule in current_time_slots

            # Начинаем с blockquote
            if is_current_slot and not in_blockquote:
                lines.append(f"<blockquote><b>⏰ {time_schedule}</b>")
                in_blockquote = True
            elif not is_current_slot and in_blockquote:
                lines.append("</blockquote>")
                in_blockquote = False
                lines.append(f"⏰ <b>{time_schedule}</b>")
            else:
                lines.append(f"⏰ <b>{time_schedule}</b>")

            # Добавляем старших дежурных
            for duty in group["seniors"]:
                duty_user = await stp_repo.employee.get_users(user_id=duty.user_id)
                duty_fullname = format_fullname(duty_user, True, True)
                lines.append(f"Дежурный - {duty_fullname}")

            # Добавляем помощников
            for duty in group["helpers"]:
                duty_user = await stp_repo.employee.get_users(user_id=duty.user_id)
                duty_fullname = format_fullname(duty_user, True, True)
                lines.append(f"Помощник - {duty_fullname}")

            # Проверяем, является ли следующий слот текущим
            next_is_current = False
            if i + 1 < len(sorted_times):
                next_time_schedule = sorted_times[i + 1]
                next_is_current = next_time_schedule in current_time_slots

            # Добавляем отступы
            if is_current_slot and not next_is_current and in_blockquote:
                lines.append("</blockquote>")
                in_blockquote = False
            elif not is_current_slot:
                lines.append("")
            elif is_current_slot and next_is_current:
                if current_slots_count <= 3:
                    lines.append("")

        return "\n".join(lines)

    def parse(self, *args, **kwargs):
        """Реализация абстрактного метода parse."""
        pass


class HeadScheduleParser(BaseParser):
    """Полностью оптимизированный парсер графика руководителей с кэшированием."""

    async def get_heads_for_date(
        self, date: datetime, division: str, stp_repo: MainRequestsRepo
    ) -> List[HeadInfo]:
        """Получает руководителей на дату (оптимизировано с кэшированием).

        Args:
            date: Дата
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Список руководителей на эту дату
        """
        duty_parser = DutyScheduleParser()
        duties = await duty_parser.get_duties_for_date(date, division, stp_repo)

        try:
            # Determine schedule file
            if division in ["НТП1", "НТП2"]:
                schedule_file = self.file_manager.find_schedule_file("НТП2")
            else:
                schedule_file = self.file_manager.find_schedule_file(division)

            if not schedule_file:
                raise FileNotFoundError(f"Schedule file for {division} not found")

            # Use FastExcelReader with caching
            reader = ExcelReader(schedule_file, "ГРАФИК")
            df = reader.df

            # Find date column
            date_col = reader.find_date_column(date)
            if date_col is None:
                logger.warning(f"Date {date.day} not found in head schedule")
                return []

            candidate_heads = []
            names_to_fetch = set()

            # Scan through rows to find heads
            for row_idx in range(len(df)):
                position_found = False
                name = ""

                # Look for position and name in first columns
                for col_idx in range(min(5, df.shape[1])):
                    cell_value = reader.get_cell(row_idx, col_idx)

                    if "Руководитель группы" in cell_value:
                        position_found = True

                    if (
                        not name
                        and len(cell_value.split()) >= 3
                        and re.search(r"[А-Яа-я]", cell_value)
                        and "Руководитель" not in cell_value
                    ):
                        name = cell_value.strip()

                if not position_found or not name:
                    continue

                # Check schedule for this date
                if date_col < df.shape[1]:
                    schedule_cell = reader.get_cell(row_idx, date_col)
                    if schedule_cell and schedule_cell.strip():
                        if self.is_time_format(schedule_cell):
                            names_to_fetch.add(name)
                            candidate_heads.append({
                                "name": name,
                                "schedule": schedule_cell.strip(),
                            })

            if not names_to_fetch:
                return []

            employee_cache = {}
            for name in names_to_fetch:
                try:
                    user: Employee = await stp_repo.employee.get_users(fullname=name)
                    if user:
                        employee_cache[name] = user
                except Exception as e:
                    logger.debug(f"Error getting user {name}: {e}")

            # OPTIMIZATION 3: Build heads list from cached data
            heads = []
            for candidate in candidate_heads:
                name = candidate["name"]
                if name not in employee_cache:
                    continue

                user = employee_cache[name]
                duty_info = await self._check_duty_for_head(name, duties)

                heads.append(
                    HeadInfo(
                        name=name,
                        user_id=user.user_id,
                        username=user.username,
                        schedule=candidate["schedule"],
                        duty_info=duty_info,
                    )
                )

            logger.info(
                f"[Optimized] Found {len(heads)} heads for {date.strftime('%d.%m.%Y')}"
            )
            return heads

        except Exception as e:
            logger.error(f"Error getting heads: {e}")
            return []

    async def _check_duty_for_head(
        self, head_name: str, duties: List[DutyInfo]
    ) -> Optional[str]:
        """Проверяет есть ли руководитель в графике дежурных.

        Args:
            head_name: ФИО руководителя
            duties: Список дежурных

        Returns:
            График дежурств если есть, иначе None
        """
        try:
            for duty in duties:
                if self.names_match(head_name, duty.name):
                    return f"{duty.schedule} {duty.shift_type}"
            return None
        except Exception as e:
            logger.debug(
                f"[График] Ошибка проверки дежурств для руководителя {head_name}: {e}"
            )
            return None

    def format_schedule(self, heads: List[HeadInfo], date: datetime) -> str:
        """Форматирует график руководителей для отображения.

        Args:
            heads: Список руководителей
            date: Дата

        Returns:
            Отформатированная строка с графиком руководителей
        """
        if not heads:
            return f"<b>👑 Руководители • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдено руководителей на эту дату"

        lines = [f"<b>👑 Руководители • {date.strftime('%d.%m.%Y')}</b>\n"]

        # Group by time
        time_groups = {}
        for head in heads:
            time_schedule = head.schedule
            if not time_schedule or not self.is_time_format(time_schedule):
                continue

            time_match = re.search(r"(\d{1,2}:\d{2}-\d{1,2}:\d{2})", time_schedule)
            time_key = time_match.group(1) if time_match else time_schedule

            if time_key not in time_groups:
                time_groups[time_key] = []
            time_groups[time_key].append(head)

        # Sort by time
        sorted_times = sorted(
            time_groups.keys(), key=lambda t: self.parse_time_range(t)[0]
        )

        for time_schedule in sorted_times:
            group_heads = time_groups[time_schedule]
            lines.append(f"⏰ <b>{time_schedule}</b>")

            for head in group_heads:
                head_line = f"{format_fullname(fullname=head.name, username=head.username, user_id=head.user_id, short=True, gender_emoji=True)}"

                if head.duty_info:
                    head_line += f" ({head.duty_info})"

                lines.append(head_line)

            lines.append("")

        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines)

    def parse(self, *args, **kwargs):
        """Реализация абстрактного метода parse."""
        pass


class GroupScheduleParser(BaseParser):
    """Полностью оптимизированный парсер группового графика с кэшированием."""

    def __init__(self, uploads_folder: str = "uploads"):
        super().__init__(uploads_folder)
        self.duty_parser = DutyScheduleParser(uploads_folder)

    def _group_members_by_start_time(
        self, members: List[GroupMemberInfo]
    ) -> Dict[str, List[GroupMemberInfo]]:
        """Группирует членов группы по времени начала работы.

        Args:
            members: Список членов группы

        Returns:
            Словарь {время_начала: список_сотрудников}
        """
        grouped = defaultdict(list)

        for member in members:
            start_time = self._extract_start_time(member.working_hours)
            grouped[start_time].append(member)

        return dict(grouped)

    @staticmethod
    def _extract_start_time(working_hours: str) -> str | None | Any:
        """Извлекает время начала из строки рабочих часов.

        Args:
            working_hours: Строка с рабочими часами

        Returns:
            Время начала работы
        """
        if not working_hours or working_hours is None:
            return None

        time_pattern = r"(\d{1,2}:\d{2})"
        match = re.search(time_pattern, working_hours)

        if match:
            return match.group(1)

        return None

    def _format_member_with_link(self, member: GroupMemberInfo) -> str:
        """Format member name with link and working hours."""
        user_link = format_fullname(
            fullname=member.name,
            username=member.username,
            user_id=member.user_id,
            short=True,
            gender_emoji=True,
        )

        working_hours = member.working_hours or None
        result = f"{user_link} <code>{working_hours}</code>"

        if member.duty_info:
            result += f" ({member.duty_info})"

        return result

    def _sort_members_by_time(
        self, members: List[GroupMemberInfo]
    ) -> List[GroupMemberInfo]:
        """Sort members by start time."""
        return sorted(
            members,
            key=lambda m: self._parse_time_for_sorting(
                self._extract_start_time(m.working_hours)
            ),
        )

    @staticmethod
    def _parse_time_for_sorting(time_str: str) -> Tuple[int, int]:
        """Parse time string for sorting purposes."""
        if not time_str or time_str == "Не указано":
            return 99, 0

        try:
            hour, minute = time_str.split(":")
            return int(hour), int(minute)
        except (ValueError, IndexError):
            return 99, 0

    async def get_group_members(
        self, head_fullname: str, date: datetime, division: str, stp_repo
    ) -> List[GroupMemberInfo]:
        """Получает членов группы для руководителя (оптимизировано с кэшированием).

        Args:
            head_fullname: ФИО руководителя
            date: Дата
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Список членов группы
        """
        try:
            group_members = []

            # Determine divisions to check
            if "НТП" in division:
                divisions_to_check = ["НТП1", "НТП2"]
            else:
                divisions_to_check = [division]

            for div in divisions_to_check:
                schedule_file = self.file_manager.find_schedule_file(div)
                if not schedule_file:
                    logger.warning(f"Schedule file for {div} not found")
                    continue

                # Use FastExcelReader with caching
                reader = ExcelReader(schedule_file, "ГРАФИК")

                # Find date column
                date_column = reader.find_date_column(date)

                # Process members from this division
                division_members = await self._process_division_members(
                    reader, head_fullname, date_column, stp_repo
                )
                group_members.extend(division_members)

            # Fetch duty information for all members
            if group_members:
                for div in divisions_to_check:
                    try:
                        duties = await self.duty_parser.get_duties_for_date(
                            date, div, stp_repo
                        )
                        for member in group_members:
                            if not hasattr(member, "duty_info") or not member.duty_info:
                                for duty in duties:
                                    if self.names_match(member.name, duty.name):
                                        member.duty_info = (
                                            f"{duty.schedule} {duty.shift_type}"
                                        )
                                        break
                    except Exception as duty_error:
                        logger.warning(
                            f"Could not fetch duty information for {div}: {duty_error}"
                        )

            logger.info(
                f"[Optimized] Found {len(group_members)} members for head {head_fullname}"
            )
            return self._sort_members_by_time(group_members)

        except Exception as e:
            logger.error(f"Error getting group members for head: {e}")
            return []

    async def _process_division_members(
        self, reader: ExcelReader, head_fullname: str, date_column, stp_repo
    ) -> List[GroupMemberInfo]:
        """Обрабатывает членов группы из файла направления.

        Args:
            reader: Читатель Excel
            head_fullname: ФИО руководителя
            date_column: Индекс столбца даты
            stp_repo: Репозиторий базы данных

        Returns:
            Список членов группы из этого направления
        """
        division_members = []
        df = reader.df

        # OPTIMIZATION 1: First pass - collect all candidate members
        candidate_members = []
        names_to_fetch = set()

        for row_idx in range(len(df)):
            name_cell = reader.get_cell(row_idx, 0)
            schedule_cell = reader.get_cell(row_idx, 1)
            position_cell = reader.get_cell(row_idx, 4)
            head_cell = reader.get_cell(row_idx, 5)

            # Check if this person belongs to the specified head
            if not self.names_match(head_fullname, head_cell):
                continue

            if not name_cell or len(name_cell.split()) < 2:
                continue

            # Get working hours for the specific date
            working_hours = None
            if date_column is not None:
                hours_cell = reader.get_cell(row_idx, date_column)
                if hours_cell and hours_cell.strip():
                    if self.is_time_format(hours_cell):
                        working_hours = hours_cell
                    else:
                        # Non-time value - skip
                        continue
                else:
                    # Empty cell - skip
                    continue

            name_stripped = name_cell.strip()
            names_to_fetch.add(name_stripped)
            candidate_members.append({
                "name": name_stripped,
                "schedule": schedule_cell.strip() if schedule_cell else None,
                "position": position_cell.strip() if position_cell else "Специалист",
                "working_hours": working_hours,
            })

        if not names_to_fetch:
            return []

        # OPTIMIZATION 2: Batch fetch all employees
        employee_cache = {}
        for name in names_to_fetch:
            try:
                user = await stp_repo.employee.get_users(fullname=name)
                if user:
                    employee_cache[name] = user
            except Exception as e:
                logger.debug(f"Error getting user {name}: {e}")

        # OPTIMIZATION 3: Build member list from cached data
        for candidate in candidate_members:
            name = candidate["name"]
            if name not in employee_cache:
                logger.debug(f"User {name} not found in DB, skipping")
                continue

            user = employee_cache[name]
            member = GroupMemberInfo(
                name=name,
                user_id=user.user_id,
                username=user.username,
                schedule=candidate["schedule"],
                position=candidate["position"],
                working_hours=candidate["working_hours"],
            )
            division_members.append(member)

        return division_members

    async def get_group_members_for_user(
        self, user_fullname: str, date: datetime, division: str, stp_repo
    ) -> List[GroupMemberInfo]:
        """Получает коллег по группе для пользователя.

        Args:
            user_fullname: ФИО пользователя
            date: Дата
            division: Направление
            stp_repo: Репозиторий базы данных

        Returns:
            Список коллег по группе
        """
        try:
            user = await stp_repo.employee.get_users(fullname=user_fullname)
            if not user or not user.head:
                logger.warning(
                    f"User {user_fullname} not found or has no head assigned"
                )
                return []

            # Get all members under the same head
            all_members = await self.get_group_members(
                user.head, date, division, stp_repo
            )

            return self._sort_members_by_time(all_members)

        except Exception as e:
            logger.error(f"Error getting colleagues for user: {e}")
            return []

    def format_group_schedule_for_head(
        self,
        date: datetime,
        group_members: List[GroupMemberInfo],
        page: int = 1,
        members_per_page: int = 20,
    ) -> Tuple[str, int, bool, bool]:
        """Форматирует групповой график для руководителя с пагинацией.

        Args:
            date: Дата
            group_members: Список членов группы
            page: Номер страницы
            members_per_page: Членов на странице

        Returns:
            Кортеж (текст, всего_страниц, есть_предыдущая, есть_следующая)
        """
        if not group_members:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ Не найдены участники группы",
                1,
                False,
                False,
            )

        # Apply pagination
        total_members = len(group_members)
        total_pages = max(1, (total_members + members_per_page - 1) // members_per_page)

        start_idx = (page - 1) * members_per_page
        end_idx = start_idx + members_per_page
        page_members = group_members[start_idx:end_idx]

        # Group by start time
        grouped_by_start_time = self._group_members_by_start_time(page_members)
        sorted_start_times = sorted(
            grouped_by_start_time.keys(), key=self._parse_time_for_sorting
        )

        # Build message
        lines = [f"❤️ <b>Твоя группа • {date.strftime('%d.%m.%Y')}</b>"]

        if total_pages > 1:
            lines.append(
                f"<i>Страница {page} из {total_pages}\nОтображено {len(page_members)} из {total_members} участников</i>"
            )

        lines.append("")

        for start_time in sorted_start_times:
            members = grouped_by_start_time[start_time]
            lines.append(f"⏰ <b>{start_time}</b>")

            for member in members:
                lines.append(self._format_member_with_link(member))

            lines.append("")

        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines), total_pages, page > 1, page < total_pages

    def format_group_schedule_for_user(
        self,
        date: datetime,
        group_members: List[GroupMemberInfo],
        page: int = 1,
        members_per_page: int = 20,
    ) -> Tuple[str, int, bool, bool]:
        """Форматирует групповой график для пользователя с пагинацией.

        Args:
            date: Дата
            group_members: Список членов группы
            page: Номер страницы
            members_per_page: Членов на странице

        Returns:
            Кортеж (текст, всего_страниц, есть_предыдущая, есть_следующая)
        """
        if not group_members:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ График для группы не найден",
                1,
                False,
                False,
            )

        colleagues = [member for member in group_members]

        if not colleagues:
            return (
                f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>\n\n❌ График для группы не найден",
                1,
                False,
                False,
            )

        # Apply pagination
        total_colleagues = len(group_members)
        total_pages = max(
            1, (total_colleagues + members_per_page - 1) // members_per_page
        )

        start_idx = (page - 1) * members_per_page
        end_idx = start_idx + members_per_page
        page_colleagues = colleagues[start_idx:end_idx]

        # Group by start time
        grouped_by_start_time = self._group_members_by_start_time(page_colleagues)
        sorted_start_times = sorted(
            grouped_by_start_time.keys(), key=self._parse_time_for_sorting
        )

        # Build message
        lines = [f"❤️ <b>Моя группа • {date.strftime('%d.%m.%Y')}</b>"]

        if total_pages > 1:
            lines.append(
                f"<i>Страница {page} из {total_pages}\nОтображено {len(page_colleagues)} из {total_colleagues} участников</i>"
            )

        lines.append("")

        for start_time in sorted_start_times:
            members = grouped_by_start_time[start_time]
            lines.append(f"⏰ <b>{start_time}</b>")

            for member in members:
                lines.append(self._format_member_with_link(member))

            lines.append("")

        if lines and lines[-1] == "":
            lines.pop()

        return "\n".join(lines), total_pages, page > 1, page < total_pages

    def parse(self, *args, **kwargs):
        """Реализация абстрактного метода parse."""
        pass
